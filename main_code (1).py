import openpyxl
import xlsxwriter
import re
import csv
from bs4 import BeautifulSoup
import nltk
from nltk.tokenize import word_tokenize, sent_tokenize
from bs4 import BeautifulSoup
import requests
import string
from nltk.corpus import stopwords
global tokenize_sent, tokenize_words, file_contents, complex_word

workbook = xlsxwriter.Workbook(r"C:\Users\shiva\BS4FOLDER\projectfile.xlsx")
worksheet = workbook.add_worksheet()
header_names = ['URL', 'URL_ID', 'POSITIVE SCORE', 'NEGATIVE SCORE', 'POLARITY SCORE', 'SUBJECTIVITY SCORE', 'AVG NUMBER OF WORDS PER SENTENCE', 'AVG WORD LENGTH', 'SYLLABLE PER WORD', 'COMPLEX WORD COUNT', 'AVG SENTENCE LENGTH', 'PERCENTAGE OF COMPLEX WORDS',
                'FOG INDEX', 'WORD COUNT', 'PERSONAL PRONOUNS', ]
for x in range(len(header_names)):
    # creating header names for the xlsx file
    worksheet.write(0, x, header_names[x])


def list_con(arg):
    with open(f'{arg}.txt', 'r') as f:
        obj = f.read()
        obj = obj.replace('\n', '|')
        obj = obj.replace('|', ' ').split()
    return obj

# creating a list of stopwords using the stopwords given in the stop_words file


lst_stopwords = []
for x in range(1, 8):
    count = list_con(f'stopwords/stop{x}')
    lst_stopwords += count

# creating lists of positive and negative words from the given master_dicitionary

pos_words_list = list_con(f'masterdict/positive-words')
neg_words_list = list_con(f'masterdict/negative-words')

# function to do the sentimental_analyis on each URL which returns the positive score, negative score ,polarity score and subjectivity score


def sentimental_analysis():
    global tokenize_words, tokenize_sent, file_contents
    with open(f'URL_ID/URL_ID.txt', 'r', encoding='utf-8') as f:
        file_contents = f.read()
        tokenize_words = word_tokenize(file_contents)
        tokenize_sent = sent_tokenize(file_contents)
        clean_lst = []
        for word in tokenize_words:
            if word not in lst_stopwords:
                clean_lst.append(word)

    pos_list = []
    neg_list = []
    for word in clean_lst:
        if word in pos_words_list:
            pos_list.append(word)
        if word in neg_words_list:
            neg_list.append(word)

    polarity = (len(pos_list) - len(neg_list)) / \
        ((len(pos_list) + len(neg_list)) + 0.000001)
    subjectivity = (len(pos_list) + len(neg_list)) / \
        ((len(clean_lst) + 0.000001))

    return len(pos_list), len(neg_list), polarity, subjectivity

# function to calculate the Average number of words per sentence and returning it


def avg():
    avg_no_words_per_sent = len(tokenize_words)/len(tokenize_sent)

    return avg_no_words_per_sent

# function to calculate the average word length and returning it


def AVG_WORD_LEN():
    avg_word_len = len(file_contents.replace(' ', '')) / len(file_contents)

    return avg_word_len

# function to calculate the number of syllables in a word


def count_syllables(word):

    num_vowels = len([v for v in word if v in 'aeiouAEIOU'])

    if word[-2:] in ['es', 'ed']:
        num_vowels -= 1
    num_syllables = max(0, num_vowels)

    return num_syllables

# function to calculate the total number of syllables


def syllable_count():
    count = 0
    for word in tokenize_words:
        num_vowels = len([v for v in word if v in 'aeiouAEIOU'])

        if word[-2:] in ['es', 'ed']:
            num_vowels -= 1
            count += 1

    return count

# function to calculate the total number of complex words


def complex():
    global complex_word
    complex_word = 0
    for word in tokenize_words:
        if count_syllables(word) > 2:
            complex_word += 1

    return complex_word

# function to do the readability analysis and return the average sentence length, percentange of complex words, and fog index


def readability_analysis():
    avg_sent_len = len(tokenize_words)/len(tokenize_sent)
    percentage_of_complex_words = complex_word/len(tokenize_words)
    fog_index = 0.4*(avg_sent_len + percentage_of_complex_words)

    return avg_sent_len, percentage_of_complex_words, fog_index

# function to calculate the word_count after removing all the stop words and punctuations from the text


def word_count():
    stop_words = stopwords.words('english')
    translator = str.maketrans('', '', string.punctuation)
    fc_punc = file_contents.translate(translator)
    fc_punc = word_tokenize(fc_punc)

    clean_file_contents = []
    for word in fc_punc:
        if word not in stop_words:
            clean_file_contents.append(word)

    return len(clean_file_contents)

# function to calculate the total number of personal pronouns in the text


def count_personal_pronouns():
    pronoun_count = re.compile(r'\b(I|we|ours|my|(?-i:us))\b', re.I)
    pronouns = pronoun_count.findall(file_contents)

    return len(pronouns)


contents = []
url_id = 37
y_axis = 1
upd = {}


# Open file in read mode by donwloading the given input file and converting into a CSV file
with open('input.csv', 'r') as csvf:
    urls = csv.reader(csvf)

    for url in urls:
        contents.append(url[1])  # Add each url to list content
for url in contents:  # Parse through each url in the list.
    try:
        soup = BeautifulSoup(requests.get(url).content, 'html.parser')
        article = soup.find('article')
        title = article.find(
            'h1', class_=['entry-title', 'tdb-title-text']).text
        div_element = article.select_one(
            '.td-post-content.tagdiv-type,.tdb_single_content .tdb-block-inner.td-fix-index').text
        with open(f'URL_ID/URL_ID.txt', 'w+', encoding='utf-8') as f:
            f.write(title)
            # writing the title and the article text scrapped from the url into a text file named URL_ID
            f.write(div_element)

        upd['url'] = url
        upd['url_id'] = url_id
        upd['pos'], upd['neg'], upd['pol'], upd['sub'] = sentimental_analysis()
        upd['avg_words_per_sent'] = avg()
        upd['avg_word_len'] = AVG_WORD_LEN()
        upd['syllable_count'] = syllable_count()
        upd['comp_words'] = complex()
        upd['avg_sent_len'], upd['per_comp_words'], upd['fog_index'] = readability_analysis()
        upd['word_count'] = word_count()
        upd['per_pronoun'] = count_personal_pronouns()

        counting = 0
        for x in upd:
            # writing all the values of each header in the xlsx file
            worksheet.write(y_axis, counting, upd[x])
            counting += 1

        y_axis += 1
        url_id += 1

    except Exception as e:
        print(e)
        # when the page is not found, the except block will leave the entire row empty
        worksheet.write(y_axis, 1, '')
        worksheet.write(y_axis, 0, '')
        y_axis += 1
        url_id += 1
        continue
workbook.close()


# deleting the rows which are empty
workbook = openpyxl.load_workbook('projectfile.xlsx')
worksheet = workbook['Sheet1']
for row in range(1, worksheet.max_row + 1):
    if all(cell.value is None for cell in worksheet[row]):

        worksheet.delete_rows(row)

workbook.save('projectfile.xlsx')
